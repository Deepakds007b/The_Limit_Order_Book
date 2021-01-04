from flask import Flask, render_template, request, redirect
from datetime import datetime
import json
import xlrd
import openpyxl
from openpyxl import load_workbook
date = datetime.now()
sell = ("C:\\Users\\vishal\\Desktop\\order book\\sell.xlsx")
buy = ("C:\\Users\\vishal\\Desktop\\order book\\buy.xlsx")
sell_price=[]
sell_share=[]
buy_price=[]
buy_share=[]
sell_price.clear()
sell_share.clear()
buy_price.clear()
buy_share.clear()
wb1 = xlrd.open_workbook(buy)
sheet1 = wb1.sheet_by_index(0)
wb2 = xlrd.open_workbook(sell)
sheet2 = wb2.sheet_by_index(0)
wb1 = load_workbook(buy)
ws1 = wb1.worksheets[0]
wb2 = load_workbook(sell)
ws2 = wb2.worksheets[0]


app = Flask(__name__)
# app.config['UPLOAD_FOLDER']= value['upload_location']


@app.route('/')
def home():
    return render_template("base.html")

@app.route('/orderbook', methods=['GET', 'POST'])
def orderbook():
    sell_price=[]
    sell_share=[]
    buy_price=[]
    buy_share=[]
    sell_price.clear()
    sell_share.clear()
    buy_price.clear()
    buy_share.clear()
    wb1 = xlrd.open_workbook(buy)
    sheet1 = wb1.sheet_by_index(0)
    wb2 = xlrd.open_workbook(sell)
    sheet2 = wb2.sheet_by_index(0)
    wb1 = load_workbook(buy)
    ws1 = wb1.worksheets[0]
    wb2 = load_workbook(sell)
    ws2 = wb2.worksheets[0]
    n = 5
    i=0
    while (i < n):
        if (sheet1.cell_value(i, 1) <= 0):
            n = n+1
        elif (sheet1.cell_value(i, 1) != 0):
            buy_price.append(sheet1.cell_value(i,0))
            buy_share.append(sheet1.cell_value(i,1))
        i = i+1
    m = 5
    j=0
    while (j < m):
        if (sheet2.cell_value(j, 1) <= 0):
            m = m+1
        elif (sheet2.cell_value(j, 1) != 0):
            sell_price.append(sheet2.cell_value(j,0))
            sell_share.append(sheet2.cell_value(j,1))
        j = j+1
    return render_template("orderbook.html", date=date, buy_price = buy_price,buy_share = buy_share,sell_price = sell_price,sell_share = sell_share)
    
@app.route('/MarketOrder', methods=['GET', 'POST'])
def MarketOrder():    
    sell_price=[]
    sell_share=[]
    buy_price=[]
    buy_share=[]
    sell_price.clear()
    sell_share.clear()
    buy_price.clear()
    buy_share.clear()
    wb1 = xlrd.open_workbook(buy)
    sheet1 = wb1.sheet_by_index(0)
    wb2 = xlrd.open_workbook(sell)
    sheet2 = wb2.sheet_by_index(0)
    n = 5
    i=0
    while (i < n):
        if (sheet1.cell_value(i, 1) <= 0):
            n = n+1
        elif (sheet1.cell_value(i, 1) != 0):
            buy_price.append(sheet1.cell_value(i,0))
            buy_share.append(sheet1.cell_value(i,1))
        i = i+1
    m = 5
    j=0
    while (j < m):
        if (sheet2.cell_value(j, 1) <= 0):
            m = m+1
        elif (sheet2.cell_value(j, 1) != 0):
            sell_price.append(sheet2.cell_value(j,0))
            sell_share.append(sheet2.cell_value(j,1))
        j = j+1
    return render_template("MarketOrder.html", date=date, buy_price = buy_price,buy_share = buy_share,sell_price = sell_price,sell_share = sell_share)

@app.route('/marketorderbuying', methods=['GET', 'POST'])
def MarketOrderbuying():
    wb1 = xlrd.open_workbook(buy)
    sheet1 = wb1.sheet_by_index(0)
    wb2 = xlrd.open_workbook(sell)
    sheet2 = wb2.sheet_by_index(0)
    wb1 = load_workbook(buy)
    ws1 = wb1.worksheets[0]
    wb2 = load_workbook(sell)
    ws2 = wb2.worksheets[0]
    if (request.method == 'POST'):
        entered_size = request.form.get('size')
        radio_value = request.form.get('rad_val', False)
        print(entered_size)
        print(radio_value)
        i=0
        if(radio_value=="Buy"):
            if (sheet1.cell_value(i, 1) < float(entered_size)):
                d = ws1.cell(row = i+1,column=2)
                entered_size = float(entered_size) - d.value
                d.value = 0   
                wb1.save(buy)    
                i = i+1
                return share(buy,float(entered_size),i)       
            elif(sheet1.cell_value(i, 1) >= float(entered_size)):
                d = ws1.cell(row=i+1, column=2)
                d.value = d.value-float(entered_size)
        i=0
        if(radio_value=="Sell"):
            if (sheet2.cell_value(i, 1) < float(entered_size)):
                d = ws2.cell(row = i+1,column=2)
                entered_size = float(entered_size) - d.value
                d.value = 0   
                wb2.save(sell)    
                i = i+1
                return share(sell,float(entered_size),i)       
            elif(sheet2.cell_value(i, 1) >= float(entered_size)):
                d = ws2.cell(row=i+1, column=2)
                d.value = d.value-float(entered_size)

            wb1.save(buy)
            wb2.save(sell)
        

    return redirect("/MarketOrder")

def share(filename, s_value, i):
    wb1 = xlrd.open_workbook(filename)
    sheet = wb1.sheet_by_index(0)
    # sheet.cell_value(0, 0)
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    if (sheet.cell_value(i, 1) < s_value):
        d = ws.cell(row = i+1,column=2)
        s_value = s_value - d.value
        d.value = 0   
        wb.save(filename)    
        i = i+1
        return share(filename,s_value,i)       
    elif(sheet.cell_value(i, 1) >= s_value):
        d = ws.cell(row=i+1, column=2)
        d.value = d.value-s_value

    wb.save(filename)
    return redirect("/MarketOrder")

@app.route('/limitorderbuying' , methods = ['GET','POST'] )
def limitorderbuying():
    wb1 = xlrd.open_workbook(buy)
    sheet1 = wb1.sheet_by_index(0)
    wb2 = xlrd.open_workbook(sell)
    sheet2 = wb2.sheet_by_index(0)
    wb1 = load_workbook(buy)
    ws1 = wb1.worksheets[0]
    wb2 = load_workbook(sell)
    ws2 = wb2.worksheets[0]
    if (request.method == 'POST'):
        entered_price = request.form.get('price')
        entered_size = request.form.get('size')
        radio_value = request.form.get('rad_val', False)
        if (radio_value=="Buy"):
            z=0
            for z in range (sheet1.nrows):
                if (sheet1.cell_value(z,0) == float(entered_price) and sheet1.cell_value(z,1)>=float(float(entered_size))):
                    d = ws1.cell(row = z+1,column=2)
                    d.value = d.value - float(float(entered_size))
                    wb1.save(buy)
            z=0
            for z in range (sheet2.nrows):
                if (sheet2.cell_value(z,0) == float(entered_price) and sheet2.cell_value(z,1)>=float(float(entered_size))):
                    d = ws2.cell(row = z+1,column=2)
                    d.value = d.value - float(float(entered_size))
                    wb2.save(sell)
        elif (radio_value=="Sell"):
            z=0
            for z in range (sheet1.nrows):
                if (sheet1.cell_value(z,0)==float(entered_price)):
                    d = ws1.cell(row = z+1,column=2)
                    print(d.value)
                    d.value = d.value + float(float(entered_size))
                    print(d.value)
                    wb1.save(buy)
            z=0
            for z in range (sheet2.nrows):
                if (sheet2.cell_value(z,0)==float(entered_price)):
                    d = ws2.cell(row = z+1,column=2)
                    print(d.value)
                    d.value = d.value + float(float(entered_size))
                    print(d.value)
                    wb2.save(sell)            
        else:
            return redirect("/LimitOrder")
    
    
    return redirect("/LimitOrder")

@app.route('/LimitOrder', methods=['GET', 'POST'])
def limitorder():
    sell_price=[]
    sell_share=[]
    buy_price=[]
    buy_share=[]
    wb1 = xlrd.open_workbook(buy)
    sheet1 = wb1.sheet_by_index(0)
    wb2 = xlrd.open_workbook(sell)
    sheet2 = wb2.sheet_by_index(0)
    sell_price.clear()
    sell_share.clear()
    buy_price.clear()
    buy_share.clear()
    n = 5
    i=0
    while (i < n):
        if (sheet1.cell_value(i, 1) <= 0):
            n = n+1
        elif (sheet1.cell_value(i, 1) != 0):
            buy_price.append(sheet1.cell_value(i,0))
            buy_share.append(sheet1.cell_value(i,1))
        i = i+1
    m = 5
    j=0
    while (j < m):
        if (sheet2.cell_value(j, 1) <= 0):
            m = m+1
        elif (sheet2.cell_value(j, 1) != 0):
            sell_price.append(sheet2.cell_value(j,0))
            sell_share.append(sheet2.cell_value(j,1))
        j = j+1

    return render_template("LimitOrder.html", date=date, buy_price = buy_price,buy_share = buy_share,sell_price = sell_price,sell_share = sell_share)


if __name__ == "__main__":
    app.run(debug=True)
